"""
Exporta resultados de um experimento do LangSmith para Excel.
Inclui: bug_report, resposta gerada, referência, scores e reasoning de cada métrica.

Uso:
    python src/export_experiment.py                      # exporta o experimento mais recente
    python src/export_experiment.py <experiment_name>    # exporta experimento específico
"""

import os
import sys
from collections import defaultdict
from pathlib import Path
from dotenv import load_dotenv
from langsmith import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()


SCORE_COLORS = {
    "high":   "E2EFDA",  # >= 0.9  verde
    "medium": "FFF2CC",  # >= 0.7  amarelo
    "low":    "FCE4D6",  # < 0.7   vermelho
}

METRICS = ["f1_score", "clarity", "precision", "helpfulness", "correctness"]


def score_color(score):
    if score is None:
        return "F2F2F2"
    if score >= 0.9:
        return SCORE_COLORS["high"]
    if score >= 0.7:
        return SCORE_COLORS["medium"]
    return SCORE_COLORS["low"]


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_cell(cell, bg_color="FFFFFF", wrap=True, bold=False):
    cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = thin_border()
    if bold:
        cell.font = Font(bold=True)


def fetch_experiment_runs(client: Client, experiment_name: str, dataset_name: str):
    """Busca os runs de um experimento e o feedback associado."""
    print(f"   Buscando runs do experimento '{experiment_name}'...")

    runs = list(client.list_runs(
        project_name=experiment_name,
        execution_order=1,
        is_root=True,
    ))
    print(f"   ✓ {len(runs)} runs encontrados")

    run_ids = [str(r.id) for r in runs]

    feedbacks = list(client.list_feedback(run_ids=run_ids))
    print(f"   ✓ {len(feedbacks)} feedbacks encontrados")

    feedback_map = defaultdict(dict)
    for fb in feedbacks:
        run_id = str(fb.run_id)
        key = fb.key
        feedback_map[run_id][key] = {
            "score": fb.score,
            "comment": fb.comment or "",
        }

    return runs, feedback_map


def list_experiments(client: Client, dataset_name: str):
    """Lista experimentos associados ao dataset."""
    try:
        sessions = list(client.list_projects(
            reference_dataset_name=dataset_name,
        ))
        return sorted(sessions, key=lambda s: s.modified_at or s.created_at, reverse=True)
    except Exception as e:
        print(f"⚠️  Erro ao listar experimentos: {e}")
        return []


def build_excel(runs, feedback_map, experiment_name: str, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Cabeçalho
    headers = [
        "#", "Bug Report", "Resposta Gerada", "Referência",
        "F1\nScore", "Clarity", "Precision", "Helpfulness", "Correctness",
        "Reasoning\nF1", "Reasoning\nClarity", "Reasoning\nPrecision",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        style_header(cell)

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    col_widths = [4, 45, 55, 55, 8, 8, 8, 10, 10, 45, 45, 45]
    for col_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Linhas de dados
    for row_idx, run in enumerate(runs, start=2):
        run_id = str(run.id)
        fb = feedback_map.get(run_id, {})

        bug_report  = (run.inputs or {}).get("bug_report", "")
        answer      = ((run.outputs or {}).get("answer") or
                       (run.outputs or {}).get("output") or "")
        reference   = (run.reference_example_inputs or {})  # pode não estar aqui

        # fallback: tenta pegar referência nos outputs se não vier separado
        if not reference:
            reference = ""

        scores = {m: fb.get(m, {}).get("score") for m in METRICS}
        reasons = {
            "f1_score":  fb.get("f1_score",  {}).get("comment", ""),
            "clarity":   fb.get("clarity",   {}).get("comment", ""),
            "precision": fb.get("precision", {}).get("comment", ""),
        }

        avg_score = None
        valid = [s for s in scores.values() if s is not None]
        if valid:
            avg_score = sum(valid) / len(valid)

        row_color = score_color(avg_score)

        values = [
            row_idx - 1,
            bug_report,
            answer,
            reference,
            scores.get("f1_score"),
            scores.get("clarity"),
            scores.get("precision"),
            scores.get("helpfulness"),
            scores.get("correctness"),
            reasons["f1_score"],
            reasons["clarity"],
            reasons["precision"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Colunas de score: cor individual por valor
            if col_idx in range(5, 10):
                bg = score_color(val) if isinstance(val, (int, float)) else row_color
            else:
                bg = row_color
            style_cell(cell, bg_color=bg)

        ws.row_dimensions[row_idx].height = max(60, min(len(bug_report) // 2, 150))

    # Aba de legenda
    ws2 = wb.create_sheet("Legenda")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 45

    ws2.cell(row=1, column=1, value="Legenda de Cores — Score").font = Font(bold=True, size=12)
    ws2.merge_cells("A1:B1")

    legend = [
        ("≥ 0.90 — Aprovado",  SCORE_COLORS["high"],   "Métrica atingiu o threshold mínimo"),
        ("≥ 0.70 — Atenção",   SCORE_COLORS["medium"], "Métrica abaixo do threshold, mas funcional"),
        ("< 0.70 — Reprovado", SCORE_COLORS["low"],    "Métrica com gap significativo a melhorar"),
    ]
    for r, (label, color, desc) in enumerate(legend, start=2):
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.fill = PatternFill(fill_type="solid", fgColor=color)
        c1.font = Font(bold=True)
        c1.border = thin_border()
        c1.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 20

        c2 = ws2.cell(row=r, column=2, value=desc)
        c2.fill = PatternFill(fill_type="solid", fgColor=color)
        c2.border = thin_border()
        c2.alignment = Alignment(vertical="center")

    wb.save(output_path)


def main():
    client = Client()
    dataset_name = os.getenv("LANGSMITH_PROJECT", "prompt-optimization-challenge-resolved") + "-eval"

    target_experiment = sys.argv[1] if len(sys.argv) > 1 else None

    print(f"\nListando experimentos do dataset '{dataset_name}'...")
    experiments = list_experiments(client, dataset_name)

    if not experiments:
        print("❌ Nenhum experimento encontrado.")
        return 1

    print(f"\nExperimentos disponíveis ({len(experiments)}):")
    for i, exp in enumerate(experiments[:10]):
        marker = "← mais recente" if i == 0 else ""
        print(f"  [{i}] {exp.name} {marker}")

    if target_experiment:
        experiment = next((e for e in experiments if target_experiment in e.name), None)
        if not experiment:
            print(f"❌ Experimento '{target_experiment}' não encontrado.")
            return 1
    else:
        experiment = experiments[0]

    print(f"\n✓ Exportando: {experiment.name}")

    runs, feedback_map = fetch_experiment_runs(client, experiment.name, dataset_name)

    if not runs:
        print("❌ Nenhum run encontrado no experimento.")
        return 1

    safe_name = experiment.name.replace("/", "_").replace(" ", "_")[:50]
    output_path = f"datasets/resultado_{safe_name}.xlsx"

    print(f"\n   Gerando Excel...")
    build_excel(runs, feedback_map, experiment.name, output_path)

    print(f"\n✅ Exportado: {output_path}")
    print(f"   {len(runs)} exemplos | métricas + reasoning por exemplo")
    return 0


if __name__ == "__main__":
    sys.exit(main())
