import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_jsonl(path: str):
    examples = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                examples.append(json.loads(line))
    return examples


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_meta(cell, color="D6E4F0"):
    cell.fill = PatternFill(fill_type="solid", fgColor=color)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


COMPLEXITY_COLORS = {
    "simple": "E2EFDA",
    "medium": "FFF2CC",
    "complex": "FCE4D6",
}


def main():
    jsonl_path = "datasets/bug_to_user_story.jsonl"
    examples = load_jsonl(jsonl_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset"

    headers = [
        "#",
        "Bug Report",
        "User Story (Referência)",
        "Domain",
        "Type",
        "Complexity",
        "Severity",
    ]

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col_idx))

    ws.row_dimensions[1].height = 30

    col_widths = [4, 55, 70, 16, 30, 12, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, example in enumerate(examples, start=2):
        bug_report = example.get("inputs", {}).get("bug_report", "")
        reference = example.get("outputs", {}).get("reference", "")
        meta = example.get("metadata", {})

        domain = meta.get("domain", "")
        bug_type = meta.get("type", "")
        complexity = meta.get("complexity", "")
        severity = meta.get("severity", "")

        row = [row_idx - 1, bug_report, reference, domain, bug_type, complexity, severity]
        ws.append(row)

        color = COMPLEXITY_COLORS.get(complexity, "FFFFFF")
        for col_idx, _ in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            if col_idx in (4, 5, 6, 7):
                style_meta(cell, color)

        ws.row_dimensions[row_idx].height = max(60, min(len(bug_report) // 2, 200))

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Legend sheet
    ws_legend = wb.create_sheet("Legenda")
    ws_legend.column_dimensions["A"].width = 20
    ws_legend.column_dimensions["B"].width = 40

    legend_title = ws_legend.cell(row=1, column=1, value="Legenda de Cores — Complexidade")
    legend_title.font = Font(bold=True, size=12)
    ws_legend.merge_cells("A1:B1")

    for row_idx, (complexity, hex_color) in enumerate(COMPLEXITY_COLORS.items(), start=2):
        label_cell = ws_legend.cell(row=row_idx, column=1, value=complexity.capitalize())
        label_cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = thin_border()
        ws_legend.row_dimensions[row_idx].height = 20

        desc = {
            "simple": "Bug isolado, 1 componente, sem impacto crítico",
            "medium": "Bug com contexto técnico, múltiplos critérios",
            "complex": "Múltiplos problemas, impacto crítico, tasks sugeridas",
        }
        desc_cell = ws_legend.cell(row=row_idx, column=2, value=desc[complexity])
        desc_cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
        desc_cell.alignment = Alignment(vertical="center")
        desc_cell.border = thin_border()

    output_path = "datasets/bug_to_user_story.xlsx"
    wb.save(output_path)
    print(f"✅ Planilha gerada: {output_path}")
    print(f"   {len(examples)} exemplos exportados")


if __name__ == "__main__":
    main()
